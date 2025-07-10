#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Integration Tests for Data Export Functionality

This module contains integration tests for the data export functionality
that exports fraud data to various formats (CSV, Excel, JSON).
"""

import os
import sys
import unittest
from unittest.mock import MagicMock, patch, ANY
from pathlib import Path
import pandas as pd
import json
import tempfile
import shutil
from datetime import datetime, timedelta

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

# Import the data export module
from utils.data_export import DataExporter


class TestDataExportIntegration(unittest.TestCase):
    """Integration tests for the data export functionality"""
    
    def setUp(self):
        """Set up test fixtures"""
        # Create a temporary directory for test exports
        self.temp_dir = tempfile.mkdtemp()
        
        # Create a mock database connection
        self.mock_db = MagicMock()
        
        # Create sample transaction data
        self.sample_transactions = pd.DataFrame({
            'transaction_id': ['T001', 'T002', 'T003', 'T004', 'T005'],
            'timestamp': [
                datetime.now() - timedelta(hours=1),
                datetime.now() - timedelta(hours=2),
                datetime.now() - timedelta(hours=3),
                datetime.now() - timedelta(hours=4),
                datetime.now() - timedelta(hours=5)
            ],
            'card_number': ['1234****5678'] * 5,
            'amount': [100.0, 200.0, 5000.0, 150.0, 6000.0],
            'merchant_id': ['M001', 'M002', 'M003', 'M001', 'M002'],
            'merchant_name': ['Merchant A', 'Merchant B', 'Merchant C', 'Merchant A', 'Merchant B'],
            'merchant_category': ['Electronics', 'Dining', 'Travel', 'Electronics', 'Dining'],
            'country': ['USA', 'Canada', 'UK', 'USA', 'Mexico'],
            'city': ['New York', 'Toronto', 'London', 'Los Angeles', 'Mexico City'],
            'latitude': [40.7128, 43.6532, 51.5074, 34.0522, 19.4326],
            'longitude': [-74.0060, -79.3832, -0.1278, -118.2437, -99.1332],
            'is_fraud': [False, False, True, False, True]
        })
        
        # Create sample fraud summary data
        self.sample_daily_summary = pd.DataFrame({
            'date': [datetime.now().date() - timedelta(days=i) for i in range(7)],
            'fraud_count': [5, 3, 7, 2, 4, 6, 3],
            'total_amount': [25000.0, 15000.0, 35000.0, 10000.0, 20000.0, 30000.0, 15000.0]
        })
        
        self.sample_category_summary = pd.DataFrame({
            'merchant_category': ['Electronics', 'Travel', 'Dining', 'Retail', 'Entertainment'],
            'fraud_count': [10, 8, 6, 4, 2],
            'total_amount': [50000.0, 40000.0, 30000.0, 20000.0, 10000.0]
        })
        
        self.sample_country_summary = pd.DataFrame({
            'country': ['USA', 'UK', 'Canada', 'Mexico', 'Brazil'],
            'fraud_count': [12, 8, 6, 4, 2],
            'total_amount': [60000.0, 40000.0, 30000.0, 20000.0, 10000.0]
        })
        
        # Create a DataExporter instance with the mock database
        with patch('utils.data_export.EXPORT_DIR', self.temp_dir):
            self.exporter = DataExporter(db_connection=self.mock_db)
    
    def tearDown(self):
        """Clean up after tests"""
        # Remove the temporary directory and its contents
        shutil.rmtree(self.temp_dir)
    
    def test_get_fraud_data_integration(self):
        """Test getting fraud transaction data from the database"""
        # Set up the mock cursor
        mock_cursor = MagicMock()
        self.mock_db.cursor.return_value.__enter__.return_value = mock_cursor
        
        # Configure the mock to return our sample data
        mock_cursor.fetchall.return_value = [
            (row.transaction_id, row.timestamp, row.card_number, row.amount,
             row.merchant_id, row.merchant_name, row.merchant_category,
             row.country, row.city, row.latitude, row.longitude, row.is_fraud)
            for _, row in self.sample_transactions[self.sample_transactions['is_fraud']].iterrows()
        ]
        mock_cursor.description = [
            ('transaction_id',), ('timestamp',), ('card_number',), ('amount',),
            ('merchant_id',), ('merchant_name',), ('merchant_category',),
            ('country',), ('city',), ('latitude',), ('longitude',), ('is_fraud',)
        ]
        
        # Call the method
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        result = self.exporter._get_fraud_data(start_date, end_date)
        
        # Verify the SQL query
        mock_cursor.execute.assert_called_once()
        query = mock_cursor.execute.call_args[0][0]
        self.assertIn('SELECT', query)
        self.assertIn('FROM fraudulent_transactions', query)
        self.assertIn('WHERE timestamp >=', query)
        self.assertIn('AND timestamp <=', query)
        
        # Verify the result
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 2)  # We have 2 fraud transactions in our sample
        self.assertTrue(all(result['is_fraud']))  # All should be fraud
    
    def test_get_fraud_summary_integration(self):
        """Test getting fraud summary data from the database"""
        # Set up the mock cursor
        mock_cursor = MagicMock()
        self.mock_db.cursor.return_value.__enter__.return_value = mock_cursor
        
        # Configure the mock to return our sample data for daily summary
        mock_cursor.fetchall.side_effect = [
            [(row.date, row.fraud_count, row.total_amount) for _, row in self.sample_daily_summary.iterrows()],
            [(row.merchant_category, row.fraud_count, row.total_amount) for _, row in self.sample_category_summary.iterrows()],
            [(row.country, row.fraud_count, row.total_amount) for _, row in self.sample_country_summary.iterrows()]
        ]
        mock_cursor.description = [
            [('date',), ('fraud_count',), ('total_amount',)],
            [('merchant_category',), ('fraud_count',), ('total_amount',)],
            [('country',), ('fraud_count',), ('total_amount',)]
        ]
        
        # Call the method
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        daily_summary, category_summary, country_summary = self.exporter._get_fraud_summary(start_date, end_date)
        
        # Verify the SQL queries
        self.assertEqual(mock_cursor.execute.call_count, 3)  # Three queries
        
        # Check the first query (daily summary)
        query1 = mock_cursor.execute.call_args_list[0][0][0]
        self.assertIn('SELECT', query1)
        self.assertIn('DATE(timestamp)', query1)
        self.assertIn('GROUP BY date', query1)
        self.assertIn('ORDER BY date', query1)
        
        # Check the second query (category summary)
        query2 = mock_cursor.execute.call_args_list[1][0][0]
        self.assertIn('SELECT', query2)
        self.assertIn('merchant_category', query2)
        self.assertIn('GROUP BY merchant_category', query2)
        self.assertIn('ORDER BY fraud_count DESC', query2)
        
        # Check the third query (country summary)
        query3 = mock_cursor.execute.call_args_list[2][0][0]
        self.assertIn('SELECT', query3)
        self.assertIn('country', query3)
        self.assertIn('GROUP BY country', query3)
        self.assertIn('ORDER BY fraud_count DESC', query3)
        
        # Verify the results
        self.assertIsInstance(daily_summary, pd.DataFrame)
        self.assertEqual(len(daily_summary), 7)  # 7 days of data
        
        self.assertIsInstance(category_summary, pd.DataFrame)
        self.assertEqual(len(category_summary), 5)  # 5 categories
        
        self.assertIsInstance(country_summary, pd.DataFrame)
        self.assertEqual(len(country_summary), 5)  # 5 countries
    
    @patch('utils.data_export.DataExporter._get_fraud_data')
    @patch('utils.data_export.DataExporter._get_fraud_summary')
    def test_export_to_csv_integration(self, mock_get_summary, mock_get_data):
        """Test exporting fraud data to CSV files"""
        # Set up the mocks
        mock_get_data.return_value = self.sample_transactions[self.sample_transactions['is_fraud']]
        mock_get_summary.return_value = (self.sample_daily_summary, self.sample_category_summary, self.sample_country_summary)
        
        # Call the method
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        result = self.exporter.export_to_csv(start_date, end_date)
        
        # Verify the mocks were called
        mock_get_data.assert_called_once_with(start_date, end_date)
        mock_get_summary.assert_called_once_with(start_date, end_date)
        
        # Verify the result
        self.assertIsInstance(result, dict)
        self.assertEqual(len(result), 4)  # 4 CSV files
        
        # Check that the files were created
        for file_path in result.values():
            self.assertTrue(os.path.exists(file_path))
            self.assertTrue(file_path.endswith('.csv'))
        
        # Check the content of the files
        # Transactions CSV
        transactions_csv = result['transactions']
        df_transactions = pd.read_csv(transactions_csv)
        self.assertEqual(len(df_transactions), 2)  # 2 fraud transactions
        
        # Daily summary CSV
        daily_csv = result['daily_summary']
        df_daily = pd.read_csv(daily_csv)
        self.assertEqual(len(df_daily), 7)  # 7 days of data
        
        # Category summary CSV
        category_csv = result['category_summary']
        df_category = pd.read_csv(category_csv)
        self.assertEqual(len(df_category), 5)  # 5 categories
        
        # Country summary CSV
        country_csv = result['country_summary']
        df_country = pd.read_csv(country_csv)
        self.assertEqual(len(df_country), 5)  # 5 countries
    
    @patch('utils.data_export.DataExporter._get_fraud_data')
    @patch('utils.data_export.DataExporter._get_fraud_summary')
    def test_export_to_excel_integration(self, mock_get_summary, mock_get_data):
        """Test exporting fraud data to an Excel file"""
        # Set up the mocks
        mock_get_data.return_value = self.sample_transactions[self.sample_transactions['is_fraud']]
        mock_get_summary.return_value = (self.sample_daily_summary, self.sample_category_summary, self.sample_country_summary)
        
        # Call the method
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        result = self.exporter.export_to_excel(start_date, end_date)
        
        # Verify the mocks were called
        mock_get_data.assert_called_once_with(start_date, end_date)
        mock_get_summary.assert_called_once_with(start_date, end_date)
        
        # Verify the result
        self.assertIsInstance(result, str)
        self.assertTrue(os.path.exists(result))
        self.assertTrue(result.endswith('.xlsx'))
        
        # Check the content of the Excel file
        import openpyxl
        wb = openpyxl.load_workbook(result)
        
        # Check that all sheets exist
        self.assertIn('Fraud Transactions', wb.sheetnames)
        self.assertIn('Daily Summary', wb.sheetnames)
        self.assertIn('Category Summary', wb.sheetnames)
        self.assertIn('Country Summary', wb.sheetnames)
        
        # Check the content of each sheet
        # Transactions sheet
        sheet_transactions = wb['Fraud Transactions']
        self.assertEqual(sheet_transactions.max_row, 3)  # Header + 2 rows
        
        # Daily summary sheet
        sheet_daily = wb['Daily Summary']
        self.assertEqual(sheet_daily.max_row, 8)  # Header + 7 rows
        
        # Category summary sheet
        sheet_category = wb['Category Summary']
        self.assertEqual(sheet_category.max_row, 6)  # Header + 5 rows
        
        # Country summary sheet
        sheet_country = wb['Country Summary']
        self.assertEqual(sheet_country.max_row, 6)  # Header + 5 rows
    
    @patch('utils.data_export.DataExporter._get_fraud_data')
    @patch('utils.data_export.DataExporter._get_fraud_summary')
    def test_export_to_json_integration(self, mock_get_summary, mock_get_data):
        """Test exporting fraud data to JSON files"""
        # Set up the mocks
        mock_get_data.return_value = self.sample_transactions[self.sample_transactions['is_fraud']]
        mock_get_summary.return_value = (self.sample_daily_summary, self.sample_category_summary, self.sample_country_summary)
        
        # Call the method
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        result = self.exporter.export_to_json(start_date, end_date)
        
        # Verify the mocks were called
        mock_get_data.assert_called_once_with(start_date, end_date)
        mock_get_summary.assert_called_once_with(start_date, end_date)
        
        # Verify the result
        self.assertIsInstance(result, dict)
        self.assertEqual(len(result), 4)  # 4 JSON files
        
        # Check that the files were created
        for file_path in result.values():
            self.assertTrue(os.path.exists(file_path))
            self.assertTrue(file_path.endswith('.json'))
        
        # Check the content of the files
        # Transactions JSON
        transactions_json = result['transactions']
        with open(transactions_json, 'r') as f:
            data_transactions = json.load(f)
        self.assertEqual(len(data_transactions), 2)  # 2 fraud transactions
        
        # Daily summary JSON
        daily_json = result['daily_summary']
        with open(daily_json, 'r') as f:
            data_daily = json.load(f)
        self.assertEqual(len(data_daily), 7)  # 7 days of data
        
        # Category summary JSON
        category_json = result['category_summary']
        with open(category_json, 'r') as f:
            data_category = json.load(f)
        self.assertEqual(len(data_category), 5)  # 5 categories
        
        # Country summary JSON
        country_json = result['country_summary']
        with open(country_json, 'r') as f:
            data_country = json.load(f)
        self.assertEqual(len(data_country), 5)  # 5 countries
    
    @patch('utils.data_export.DataExporter._get_fraud_data')
    @patch('utils.data_export.DataExporter._get_fraud_summary')
    def test_export_all_formats_integration(self, mock_get_summary, mock_get_data):
        """Test exporting fraud data to all formats at once"""
        # Set up the mocks
        mock_get_data.return_value = self.sample_transactions[self.sample_transactions['is_fraud']]
        mock_get_summary.return_value = (self.sample_daily_summary, self.sample_category_summary, self.sample_country_summary)
        
        # Call the export methods for all formats
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        
        csv_result = self.exporter.export_to_csv(start_date, end_date)
        excel_result = self.exporter.export_to_excel(start_date, end_date)
        json_result = self.exporter.export_to_json(start_date, end_date)
        
        # Verify the mocks were called
        self.assertEqual(mock_get_data.call_count, 3)  # Called for each format
        self.assertEqual(mock_get_summary.call_count, 3)  # Called for each format
        
        # Verify all files were created
        # CSV files
        for file_path in csv_result.values():
            self.assertTrue(os.path.exists(file_path))
            self.assertTrue(file_path.endswith('.csv'))
        
        # Excel file
        self.assertTrue(os.path.exists(excel_result))
        self.assertTrue(excel_result.endswith('.xlsx'))
        
        # JSON files
        for file_path in json_result.values():
            self.assertTrue(os.path.exists(file_path))
            self.assertTrue(file_path.endswith('.json'))
    
    @patch('utils.data_export.DataExporter._get_fraud_data')
    @patch('utils.data_export.DataExporter._get_fraud_summary')
    def test_export_with_custom_filename(self, mock_get_summary, mock_get_data):
        """Test exporting fraud data with custom filenames"""
        # Set up the mocks
        mock_get_data.return_value = self.sample_transactions[self.sample_transactions['is_fraud']]
        mock_get_summary.return_value = (self.sample_daily_summary, self.sample_category_summary, self.sample_country_summary)
        
        # Call the methods with custom filenames
        start_date = datetime.now().date() - timedelta(days=7)
        end_date = datetime.now().date()
        
        custom_prefix = "custom_export_"
        
        csv_result = self.exporter.export_to_csv(start_date, end_date, filename_prefix=custom_prefix)
        excel_result = self.exporter.export_to_excel(start_date, end_date, filename=f"{custom_prefix}excel_export.xlsx")
        json_result = self.exporter.export_to_json(start_date, end_date, filename_prefix=custom_prefix)
        
        # Verify the custom filenames were used
        # CSV files
        for file_path in csv_result.values():
            filename = os.path.basename(file_path)
            self.assertTrue(filename.startswith(custom_prefix))
        
        # Excel file
        excel_filename = os.path.basename(excel_result)
        self.assertEqual(excel_filename, f"{custom_prefix}excel_export.xlsx")
        
        # JSON files
        for file_path in json_result.values():
            filename = os.path.basename(file_path)
            self.assertTrue(filename.startswith(custom_prefix))


# Main test runner
if __name__ == '__main__':
    unittest.main()